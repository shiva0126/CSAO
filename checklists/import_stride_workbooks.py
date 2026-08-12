#!/usr/bin/env python3

"""
===============================================================================
Cloud Security Assessment Orchestrator (CSAO)

STRIDE Workbook Importer

Imports the manager's real assessment checklist + threat scenario workbooks
into checklists/checklist.yaml and data/attack_path_catalog.yaml.

Source workbooks (2 files, provided out of band - not committed here):

  1. STRIDE_CHECKLIST_Priority_Evidence_Map.xlsx
       - "Master STRIDE–Checklist Map"     177 checks: domain/area/item/
                                            STRIDE category/priority/severity/
                                            evidence type/pass & fail criteria
       - "Evidence & Output Format"        + collection method, output
                                            document, AP (attack path) code
       - "STRIDE Checklist Post-Val Log"   + MITRE ATT&CK mapping text and
                                            per-tactic Y/blank columns

  2. STRIDE_Threat_Scenarios_Refined.xlsx
       - "Refined Threat Scenarios"        23 scenarios, each listing which
                                            checklist IDs it validates and
                                            which named attack path it links
       - "Attack Path Linkage"             15 named attack paths (AP-01..15)
                                            with kill chain summaries

Automated tool bindings
------------------------
Only 63 of the 177 checks fall in the "Cloud Security (AWS & GCP)" domain -
the rest span AD, on-prem infrastructure, network devices, vulnerability
scanning, etc., which this framework's AWS-only evidence providers cannot
validate. Within those 63, AUTOMATED_SOURCES below is a hand-reviewed
mapping to this repo's actual Steampipe queries / Cloudsplaining finding
types / Prowler checks - populated ONLY where a specific existing query
unambiguously validates that exact control. Everything else imports with
`sources: []` and `manual_validation_required: true` rather than guessing -
see the accompanying capability-gap notes in checklists/README.md.

Usage:
  python3 checklists/import_stride_workbooks.py \
      --checklist-workbook /path/to/STRIDE_CHECKLIST_Priority_Evidence_Map.xlsx \
      --scenarios-workbook /path/to/STRIDE_Threat_Scenarios_Refined.xlsx \
      --checklist-output checklists/checklist.yaml \
      --catalog-output data/attack_path_catalog.yaml
===============================================================================
"""

import argparse
import json
import re
from collections import defaultdict
from pathlib import Path

import openpyxl
import yaml


CHECK_ID_PATTERN = re.compile(r"^\d+(\.\d+)+$")
AP_CODE_PATTERN = re.compile(r"^AP-\d+$")
SCENARIO_ID_PATTERN = re.compile(r"^D\d+$")
THREAT_NUMBER_PATTERN = re.compile(r"#?(\d+)")

TACTIC_COLUMNS = (
    "Initial Access", "Execution", "Persistence", "Privilege Escal.",
    "Defense Evasion", "Credential Access", "Discovery",
    "Lateral Movement", "Collection", "Exfiltration", "Impact",
)

# The source workbook truncates this one column header - map to the
# canonical MITRE ATT&CK tactic name for anything that gets displayed.
TACTIC_CANONICAL_NAME = {
    "Privilege Escal.": "Privilege Escalation",
}

TECHNIQUE_PATTERN = re.compile(r"(T\d{4}(?:\.\d{3})?)\s*([^;]*)")

# Hand-reviewed: checklist id -> [{tool, query_ref}, ...]. Only the "Cloud
# Security (AWS & GCP)" domain items that a specific existing query in
# queries/aws/**/*.sql, or a specific Cloudsplaining finding type,
# unambiguously validates. See module docstring.
AUTOMATED_SOURCES = {

    "1.3.1": [{"tool": "steampipe", "query_ref": "aws/ec2/public_instances"}],
    "1.3.2": [
        {"tool": "steampipe", "query_ref": "aws/vpc/ssh_open_world"},
        {"tool": "steampipe", "query_ref": "aws/vpc/rdp_open_world"},
    ],
    "1.3.3": [{"tool": "steampipe", "query_ref": "aws/ec2/missing_imdsv2"}],

    "1.4.1": [{"tool": "steampipe", "query_ref": "aws/s3/bucket_public"}],
    "1.4.2": [{"tool": "steampipe", "query_ref": "aws/s3/bucket_public"}],
    "1.4.3": [{"tool": "steampipe", "query_ref": "aws/s3/bucket_without_encryption"}],
    "1.4.5": [
        {"tool": "steampipe", "query_ref": "aws/s3/bucket_without_versioning"},
        {"tool": "steampipe", "query_ref": "aws/s3/bucket_without_object_lock"},
    ],

    "1.5.1": [{"tool": "steampipe", "query_ref": "aws/vpc/public_subnets"}],
    "1.5.2": [
        {"tool": "steampipe", "query_ref": "aws/ec2/open_security_groups"},
        {"tool": "steampipe", "query_ref": "aws/vpc/unrestricted_ingress"},
    ],
    "1.5.3": [{"tool": "steampipe", "query_ref": "aws/vpc/unrestricted_egress"}],
    "1.5.5": [{"tool": "steampipe", "query_ref": "aws/vpc/default_security_group"}],
    "1.5.7": [{"tool": "steampipe", "query_ref": "aws/vpc/flow_logs_disabled"}],

    "1.6.2": [{"tool": "steampipe", "query_ref": "aws/iam/access_keys_old"}],
    "1.6.3": [
        {"tool": "steampipe", "query_ref": "aws/iam/wildcard_permissions"},
        {"tool": "cloudsplaining", "query_ref": "ServiceWildcard"},
    ],
    "1.6.7": [{"tool": "cloudsplaining", "query_ref": "PrivilegeEscalation"}],

    "1.7.1": [{"tool": "steampipe", "query_ref": "aws/cloudtrail/multi_region_trail"}],
    "1.7.3": [{"tool": "steampipe", "query_ref": "aws/cloudtrail/log_validation"}],
    "1.7.4": [
        {"tool": "steampipe", "query_ref": "aws/s3/bucket_without_mfa_delete"},
        {"tool": "steampipe", "query_ref": "aws/s3/bucket_public"},
    ],
    "1.7.7": [{"tool": "steampipe", "query_ref": "aws/cloudtrail/trail_status"}],

}

# Items with a real source binding above, but where the query is only a
# partial signal for what the checklist item actually asks (e.g. "VPC
# architecture & subnet segmentation" is far broader than "are any subnets
# set to auto-assign public IPs") - always route these to analyst review
# even though automated evidence feeds in.
PARTIAL_AUTOMATION_STILL_MANUAL = {"1.5.1", "1.7.7"}

COLLECTOR_SUPPORT_RULES = {
    "aws_inventory": {
        "prefixes": ("1.3", "1.4", "1.5", "1.6", "1.9"),
        "mode": "supporting",
        "automatic": False,
    },
    "prowler": {
        "prefixes": ("1.3", "1.4", "1.5", "1.7", "1.9", "8.1"),
        "mode": "supporting",
        "automatic": False,
    },
    "cloudsplaining": {
        "prefixes": ("1.3.4", "1.6", "2.2", "2.8.3"),
        "mode": "automatic_or_supporting",
        "automatic": False,
    },
    "access_analyzer": {
        "explicit": ("1.1.3", "1.4.2", "1.6.5", "2.2.2", "2.2.4"),
        "mode": "supporting",
        "automatic": False,
    },
    "steampipe": {
        "prefixes": ("1.3", "1.4", "1.5", "1.6", "1.7"),
        "mode": "automatic_or_supporting",
        "automatic": False,
    },
    "config_aggregator": {
        "prefixes": ("1.5", "1.7", "8.1"),
        "mode": "supporting",
        "automatic": False,
    },
    "cartography": {
        "prefixes": ("1.3", "1.5", "1.6", "2.2", "4."),
        "mode": "supporting",
        "automatic": False,
    },
    "tenable": {
        "prefixes": ("5.",),
        "explicit": ("1.3.5", "1.9.4", "1.9.6"),
        "mode": "supporting",
        "automatic": False,
    },
}

DEFAULT_REPORT_SECTIONS = (
    "index",
    "executive_summary",
    "analyst_traceability",
    "technical_findings",
    "mitre_matrix",
    "attack_path_summary",
    "checklist_coverage",
    "assessment_coverage",
    "evidence_index",
    "assessment_metadata",
    "permission_report",
    "capability_report",
    "read_only_verification",
    "risk_summary",
    "recommendations",
    "appendices",
)


###############################################################################
# Workbook helpers
###############################################################################

def clean_key(key):

    if key is None:

        return key

    return " ".join(str(key).split())


def rows_as_dicts(ws, header_row_index, skip_leading_column=False, id_pattern=CHECK_ID_PATTERN):

    rows = list(ws.iter_rows(values_only=True))

    header = rows[header_row_index]

    if skip_leading_column:

        header = header[1:]

    header = [clean_key(h) for h in header]

    records = []

    for row in rows[header_row_index + 1:]:

        if skip_leading_column:

            row = row[1:]

        if not row or not id_pattern.match(str(row[0] or "")):

            continue

        records.append(dict(zip(header, row)))

    return records


def parse_mitre_techniques(text):

    if not text:

        return []

    return [match.group(1) for match in TECHNIQUE_PATTERN.finditer(str(text))]


def parse_tactics(post_val_row):

    return [
        TACTIC_CANONICAL_NAME.get(column, column) for column in TACTIC_COLUMNS
        if str(post_val_row.get(column) or "").strip().upper() == "Y"
    ]


def split_pipe_list(text):

    if not text:

        return []

    return [item.strip() for item in str(text).split("|") if item.strip()]


def parse_threat_numbers(value):

    if not value or value == "—":
        return []

    return [
        int(match.group(1))
        for match in THREAT_NUMBER_PATTERN.finditer(str(value))
    ]


###############################################################################
# Checklist import
###############################################################################

def build_scenario_index(scenario_rows):

    index = defaultdict(list)

    for row in scenario_rows:

        for checklist_id in split_pipe_list(row.get("Checklist IDs Validated")):

            index[checklist_id].append(row)

    return index


def build_checklist(checklist_workbook_path, scenarios_workbook_path):

    checklist_wb = openpyxl.load_workbook(checklist_workbook_path, data_only=True)
    scenarios_wb = openpyxl.load_workbook(scenarios_workbook_path, data_only=True)

    master_rows = rows_as_dicts(checklist_wb["Master STRIDE–Checklist Map"], 2)
    evidence_rows = rows_as_dicts(checklist_wb["Evidence & Output Format"], 2)
    postval_rows = rows_as_dicts(
        checklist_wb["STRIDE Checklist Post-Val Log"], 2, skip_leading_column=True
    )
    scenario_rows = rows_as_dicts(
        scenarios_wb["Refined Threat Scenarios"], 2, id_pattern=SCENARIO_ID_PATTERN
    )
    validation_rows = rows_as_dicts(
        scenarios_wb["Checklist Validation Matrix"], 1, id_pattern=re.compile(r"^\d+$")
    )

    evidence_by_id = {r["Check No."]: r for r in evidence_rows}
    postval_by_id = {r["Check No."]: r for r in postval_rows}
    scenarios_by_id = build_scenario_index(scenario_rows)
    matrix_by_check_id = defaultdict(list)
    for row in validation_rows:
        for checklist_id in split_pipe_list(row.get("Primary Check IDs")):
            matrix_by_check_id[checklist_id].append(row)

    checklist = []

    automated_count = 0

    for row in master_rows:

        check_id = row["Check No."]

        evidence = evidence_by_id.get(check_id, {})
        postval = postval_by_id.get(check_id, {})
        scenarios = scenarios_by_id.get(check_id, [])

        tactics = parse_tactics(postval)

        techniques = parse_mitre_techniques(postval.get("MITRE ATT&CK Mapping"))

        mitre_mapping = (
            {"tactics": tactics, "techniques": techniques}
            if (tactics or techniques) else None
        )

        threat_scenario = " || ".join(
            f"{s.get('Attacker Action (What They Do)', '')}: "
            f"{s.get('Exploit Path (How They Do It)', '')}"
            for s in scenarios
        ) or None

        ap_codes = []

        if evidence.get("AP Code"):

            ap_codes.append(evidence["AP Code"])

        for s in scenarios:

            code = s.get("Attack Path Linked")

            if code and code not in ap_codes:

                ap_codes.append(code)

        sources = AUTOMATED_SOURCES.get(check_id, [])

        manual_required = (
            check_id not in AUTOMATED_SOURCES
            or check_id in PARTIAL_AUTOMATION_STILL_MANUAL
        )

        if sources and not manual_required:

            automated_count += 1

        checklist.append({

            "id": check_id,
            "domain": row.get("Domain"),
            "check_area": row.get("Check Area"),
            "title": row.get("Check Item"),
            "stride_category": row.get("STRIDE Category"),
            "linked_threat_number": (
                row.get("Linked Threat #")
                if row.get("Linked Threat #") not in (None, "—") else None
            ),
            "linked_threat_numbers": parse_threat_numbers(row.get("Linked Threat #")),
            "primary_asset_at_risk": row.get("Primary Asset at Risk"),
            "priority_tier": row.get("Priority Tier"),
            "priority_score": row.get("Priority Score"),
            "severity": row.get("Severity"),
            "exploitability": row.get("Exploitability"),
            "attack_phase": row.get("Attack Phase"),
            "blast_radius": row.get("Blast Radius"),
            "chain_potential": row.get("Chain Potential"),
            "evidence_type": row.get("Expected Evidence Type"),
            "evidence_format": row.get("Evidence Format"),
            "pass_criteria": row.get("Pass Criteria (Expected Output)"),
            "fail_indicator": row.get("Fail Indicator (Red Flag)"),
            "collection_method": evidence.get("Collection Method"),
            "output_document": evidence.get("Output Document"),
            "related_attack_paths": ", ".join(ap_codes) or None,
            "threat_scenario": threat_scenario,
            "threat_scenario_ids": [
                f"THREAT-{int(m.get('Threat #')):02d}"
                for m in matrix_by_check_id.get(check_id, [])
                if m.get("Threat #")
            ],
            "mitre_mapping_raw": postval.get("MITRE ATT&CK Mapping"),
            "mitre_mapping": mitre_mapping,
            "sources": sources,
            "manual_validation_required": manual_required,

        })

    return checklist, automated_count


def build_threat_scenarios(scenarios_workbook_path):

    wb = openpyxl.load_workbook(scenarios_workbook_path, data_only=True)

    refined_rows = rows_as_dicts(
        wb["Refined Threat Scenarios"], 2, id_pattern=SCENARIO_ID_PATTERN
    )
    post_val_rows = rows_as_dicts(
        wb["STRIDE Post Validation Log"], 2, skip_leading_column=True, id_pattern=SCENARIO_ID_PATTERN
    )
    validation_rows = rows_as_dicts(
        wb["Checklist Validation Matrix"], 1, id_pattern=re.compile(r"^\d+$")
    )

    scenarios = []

    for index, (refined, validation) in enumerate(
        zip(refined_rows, validation_rows), start=1
    ):
        post_val = post_val_rows[index - 1] if index - 1 < len(post_val_rows) else {}
        mitre_techniques = parse_mitre_techniques(
            post_val.get("MITRE ATT&CK Mapping") or validation.get("MITRE ATT&CK Mapping")
        )
        mitre_tactics = parse_tactics(post_val)

        asset = refined.get("Asset /\nComponent", "")
        action = refined.get("Attacker Action\n(What They Do)", "")
        exploit = refined.get("Exploit Path\n(How They Do It)", "")
        title = " | ".join(part for part in (asset, action) if part)

        scenarios.append({
            "scenario_id": f"THREAT-{index:02d}",
            "threat_id": refined.get("#"),
            "threat_number": int(validation.get("Threat #")),
            "domain": refined.get("Domain"),
            "title": title,
            "asset_component": asset,
            "attacker_action": action,
            "exploit_path": refined.get("Exploit Path\n(How They Do It)"),
            "data_flow": refined.get("Data Flow\n(What Moves)"),
            "trust_boundary": refined.get("Trust Boundary\nCrossed"),
            "attack_objective": refined.get("Attack Objective\n(What They Want)"),
            "preconditions": refined.get("Precondition\n/ Weakness"),
            "business_impact": refined.get("Impact if\nSuccessful"),
            "impact": refined.get("Impact if\nSuccessful"),
            "risk_level": refined.get("Risk\nLevel") or validation.get("Risk"),
            "checklist_ids": split_pipe_list(refined.get("Checklist IDs\nValidated")),
            "primary_check_ids": split_pipe_list(validation.get("Primary Check IDs")),
            "attack_path_refs": split_pipe_list(refined.get("Attack Path\nLinked"))
            or split_pipe_list(validation.get("Attack Path")),
            "existing_controls": refined.get("Existing\nControl"),
            "remediation": refined.get("Remediation\nAction"),
            "mitre_mapping_raw": post_val.get("MITRE ATT&CK Mapping"),
            "mitre_tactics": mitre_tactics,
            "mitre_techniques": mitre_techniques,
            "stride_category": refined.get("STRIDE\nCategory"),
            "validation_summary": validation.get("What the Checklist Validates"),
            "exploitability": validation.get("Exploitability"),
            "attack_phase": validation.get("Attack Phase"),
            "initial_access_vector": validation.get("Initial Access\nVector"),
            "control_bypass_possible": validation.get("Control Bypass\nPossible"),
            "supporting_evidence": [
                validation.get("What the Checklist Validates"),
                refined.get("Existing\nControl"),
            ],
            "report_sections": list(DEFAULT_REPORT_SECTIONS),
        })

    return scenarios


###############################################################################
# Attack path catalog import
###############################################################################

def build_attack_path_catalog(scenarios_workbook_path):

    wb = openpyxl.load_workbook(scenarios_workbook_path, data_only=True)

    # This sheet has a title row but no subtitle row, so its header sits at
    # index 1 (every other sheet imported here has title+subtitle+header).
    rows = rows_as_dicts(wb["Attack Path Linkage"], 1, id_pattern=AP_CODE_PATTERN)

    catalog = []

    for row in rows:

        catalog.append({

            "ap_code": row.get("AP Code"),
            "name": row.get("Attack Path Name"),
            "kill_chain_summary": row.get("Kill Chain Summary"),
            "domains_spanned": row.get("Domains Spanned"),
            "stride_categories": row.get("STRIDE Categories"),
            "linked_threat_numbers": row.get("Linked Threat #s (this workbook)"),
            "linked_threat_number_list": parse_threat_numbers(
                row.get("Linked Threat #s (this workbook)")
            ),
            "entry_check_ids": split_pipe_list(row.get("Entry Check IDs")),
            "escalation_check_ids": split_pipe_list(row.get("Escalation Check IDs")),
            "detection_check_ids": split_pipe_list(row.get("Detection Check IDs")),
            "severity": row.get("Severity"),

        })

    return catalog


def control_matches_rule(control_id, rule):

    for value in rule.get("explicit", ()):
        if control_id == value:
            return True

    for prefix in rule.get("prefixes", ()):
        if control_id.startswith(prefix):
            return True

    return False


def build_collector_catalog(checklist):

    automatic_by_tool = defaultdict(list)

    for item in checklist:
        for source in item.get("sources", []):
            automatic_by_tool[source["tool"]].append(item["id"])

    collectors = []

    for collector, rule in COLLECTOR_SUPPORT_RULES.items():
        supported = [
            item["id"]
            for item in checklist
            if control_matches_rule(item["id"], rule)
        ]
        automatic = sorted(set(automatic_by_tool.get(collector, [])))
        manual = sorted(set(supported) - set(automatic))
        if not supported and not automatic:
            supported = []
        collectors.append({
            "collector": collector,
            "mode": rule.get("mode", "supporting"),
            "automatic_controls": automatic,
            "manual_controls": manual,
            "supported_controls": sorted(set(supported) | set(automatic)),
        })

    return collectors


def build_recommendation_catalog(checklist, scenarios):

    scenarios_by_control = defaultdict(list)
    for scenario in scenarios:
        for checklist_id in scenario.get("checklist_ids", []) or scenario.get("primary_check_ids", []):
            scenarios_by_control[checklist_id].append(scenario)

    recommendations = []

    for item in checklist:
        scenario_refs = scenarios_by_control.get(item["id"], [])
        recommendations.append({
            "recommendation_id": f"REC-TPL-{item['id']}",
            "control_id": item["id"],
            "title": f"Remediate checklist item {item['id']}: {item['title']}",
            "priority": item.get("priority_tier", ""),
            "severity": item.get("severity", ""),
            "threat_scenario_refs": [scenario["scenario_id"] for scenario in scenario_refs],
            "attack_path_refs": list(
                dict.fromkeys(
                    attack_path
                    for scenario in scenario_refs
                    for attack_path in scenario.get("attack_path_refs", [])
                )
            ) or split_pipe_list(item.get("related_attack_paths")),
            "evidence_type": item.get("evidence_type", ""),
            "evidence_format": item.get("evidence_format", ""),
            "remediation": (
                " | ".join(
                    scenario.get("remediation", "")
                    for scenario in scenario_refs
                    if scenario.get("remediation")
                )
                or item.get("output_document", "")
            ),
            "verification_steps": item.get("pass_criteria", ""),
            "report_sections": list(DEFAULT_REPORT_SECTIONS),
            "output_document": item.get("output_document", ""),
        })

    return recommendations


def build_knowledge_catalog(
    checklist_workbook_path,
    scenarios_workbook_path,
    checklist,
    attack_paths,
):

    scenarios = build_threat_scenarios(scenarios_workbook_path)
    collectors = build_collector_catalog(checklist)
    recommendations = build_recommendation_catalog(checklist, scenarios)

    scenario_ids_by_threat_number = defaultdict(list)
    for scenario in scenarios:
        if scenario.get("threat_number"):
            scenario_ids_by_threat_number[scenario["threat_number"]].append(
                scenario["scenario_id"]
            )

    for item in checklist:
        threat_numbers = item.get("linked_threat_numbers", [])
        scenario_ids = list(
            dict.fromkeys(
                scenario_id
                for threat_number in threat_numbers
                for scenario_id in scenario_ids_by_threat_number.get(threat_number, [])
            )
        )
        item["threat_scenario_ids"] = list(
            dict.fromkeys(list(item.get("threat_scenario_ids", [])) + scenario_ids)
        )

    return {
        "metadata": {
            "checklist_workbook": str(checklist_workbook_path),
            "scenarios_workbook": str(scenarios_workbook_path),
            "control_count": len(checklist),
            "scenario_count": len(scenarios),
            "attack_path_count": len(attack_paths),
        },
        "controls": checklist,
        "threat_scenarios": scenarios,
        "attack_paths": attack_paths,
        "collectors": collectors,
        "recommendations": recommendations,
        "report_sections": [{"id": item, "title": item.replace("_", " ").title()} for item in DEFAULT_REPORT_SECTIONS],
    }


###############################################################################
# Main
###############################################################################

def main():

    parser = argparse.ArgumentParser(description=__doc__)

    parser.add_argument("--checklist-workbook", required=True)
    parser.add_argument("--scenarios-workbook", required=True)
    parser.add_argument("--checklist-output", default="checklists/checklist.yaml")
    parser.add_argument("--catalog-output", default="data/attack_path_catalog.yaml")
    parser.add_argument("--knowledge-output", default="data/assessment_knowledge_catalog.yaml")

    args = parser.parse_args()

    checklist, automated_count = build_checklist(
        args.checklist_workbook, args.scenarios_workbook
    )

    catalog = build_attack_path_catalog(args.scenarios_workbook)
    knowledge_catalog = build_knowledge_catalog(
        args.checklist_workbook,
        args.scenarios_workbook,
        checklist,
        catalog,
    )

    checklist_output = Path(args.checklist_output)
    checklist_output.parent.mkdir(parents=True, exist_ok=True)

    with open(checklist_output, "w", encoding="utf-8") as f:

        yaml.safe_dump({"checklist": checklist}, f, sort_keys=False, allow_unicode=True)

    catalog_output = Path(args.catalog_output)
    catalog_output.parent.mkdir(parents=True, exist_ok=True)

    with open(catalog_output, "w", encoding="utf-8") as f:

        yaml.safe_dump({"attack_paths": catalog}, f, sort_keys=False, allow_unicode=True)

    knowledge_output = Path(args.knowledge_output)
    knowledge_output.parent.mkdir(parents=True, exist_ok=True)

    with open(knowledge_output, "w", encoding="utf-8") as f:
        yaml.safe_dump(knowledge_catalog, f, sort_keys=False, allow_unicode=True)

    print(f"Imported {len(checklist)} checklist items -> {checklist_output}")
    print(f"  {automated_count} item(s) have automated (tool, query_ref) source bindings")
    print(f"  {len(checklist) - automated_count} item(s) require manual validation / are not yet automatable")
    print(f"Imported {len(catalog)} named attack path(s) -> {catalog_output}")
    print(
        f"Imported structured knowledge catalog with "
        f"{knowledge_catalog['metadata']['scenario_count']} threat scenario(s) -> {knowledge_output}"
    )


if __name__ == "__main__":

    main()
